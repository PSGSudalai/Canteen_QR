import openpyxl
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect
from BASE.models import Transaction, PreviousOrders
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseServerError
from django.utils.dateparse import parse_date
from BASE.choices import PAYMENT_TYPE


@login_required
def generate_payment_report_all(request):
    try:
        transactions = Transaction.objects.all()
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        payment_type = request.GET.get("payment_type")

        if start_date:
            transactions = transactions.filter(
                created_at__date__gte=parse_date(start_date)
            )
        if end_date:
            transactions = transactions.filter(
                created_at__date__lte=parse_date(end_date)
            )
        if payment_type:
            transactions = transactions.filter(payment_type=payment_type)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        columns = [
            "Date",
            "Time",
            "User",
            "Amount(₹)",
            "Staff",
            "Payment Type",
            "Payment Method",
        ]
        for col_num, column_title in enumerate(columns, 1):
            column_letter = get_column_letter(col_num)
            ws[f"{column_letter}1"] = column_title

        # Populate the data
        for row_num, transaction in enumerate(transactions, 2):
            ws[f"A{row_num}"] = transaction.created_at.strftime("%Y-%m-%d")
            ws[f"B{row_num}"] = transaction.created_at.strftime("%H:%M")
            ws[
                f"C{row_num}"
            ] = f"{transaction.user.first_name} {transaction.user.last_name}".capitalize()
            ws[f"D{row_num}"] = f"₹ {transaction.amount}"
            ws[
                f"E{row_num}"
            ] = f"{transaction.staff.first_name} {transaction.staff.last_name}".capitalize()
            ws[f"F{row_num}"] = transaction.payment_type
            ws[f"G{row_num}"] = transaction.payment_method

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename=transactions_report_{payment_type}.xlsx"
            if payment_type
            else "attachment; filename=transactions_report.xlsx"
        )
        wb.save(response)
        return response
    except Exception as e:
        return HttpResponseServerError(
            f"An error occurred while generating the report: {e}"
        )


def redirect_transaction_report_page(request):
    return render(request, "reports/generate_transaction_report.html")


@login_required
def generate_product_sales_month_based_report_all(request):
    try:
        orders = PreviousOrders.objects.all()
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        if start_date:
            orders = orders.filter(created_at__date__gte=(parse_date(start_date)))
        if end_date:
            orders = orders.filter(created_at__date__lte=(parse_date(end_date)))

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Month-wise Sales Report"

        monthly_sales_data = {}
        months = set()
        total_earnings_per_month = {}

        for order in orders:
            item_name = order.item_name.capitalize()
            item_price = order.item_price
            month = order.created_at.strftime("%B").capitalize()

            if item_name not in monthly_sales_data:
                monthly_sales_data[item_name] = {"price": item_price}

            if month not in monthly_sales_data[item_name]:
                monthly_sales_data[item_name][month] = {"sales": 0, "earnings": 0}

            monthly_sales_data[item_name][month]["sales"] += order.quantity
            monthly_sales_data[item_name][month]["earnings"] += order.total

            if month not in total_earnings_per_month:
                total_earnings_per_month[month] = 0
            total_earnings_per_month[month] += order.total

            months.add(month)

        columns = ["Items", "Price(₹)"]
        for month in sorted(months):
            columns.append(f"{month} Count")
            columns.append(f"{month} Earnings")

        worksheet.append(columns)

        for item_name, sales_data in monthly_sales_data.items():
            row = [item_name, sales_data["price"]]
            for month in sorted(months):
                if month in sales_data:
                    row.append(sales_data[month]["sales"])
                    row.append(sales_data[month]["earnings"])
                else:
                    row.append(0)
                    row.append(0)

            worksheet.append(row)

        total_row = ["Total", ""]
        for month in sorted(months):
            total_row.append("")
            total_row.append(total_earnings_per_month[month])
        worksheet.append(total_row)

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response[
            "Content-Disposition"
        ] = f"attachment; filename=monthly_sales_report.xlsx"
        workbook.save(response)
        return response
    except Exception as e:
        return HttpResponseServerError(
            f"An error occurred while generating the report: {e}"
        )


@login_required
def generate_product_sales_day_based_report_all(request):
    try:
        orders = PreviousOrders.objects.all()
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        if start_date:
            orders = orders.filter(created_at__date__gte=parse_date(start_date))
        if end_date:
            orders = orders.filter(created_at__date__lte=(parse_date(end_date)))

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Day-wise Sales Report"

        daily_sales_data = {}
        days = [
            "Monday",
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday",
            "Saturday",
            "Sunday",
        ]
        days_with_data = set()
        total_earnings_per_day = {day: 0 for day in days}

        for order in orders:
            item_name = order.item_name.capitalize()
            item_price = order.item_price
            day = order.created_at.strftime("%A")

            if item_name not in daily_sales_data:
                daily_sales_data[item_name] = {"price": item_price}

            if day not in daily_sales_data[item_name]:
                daily_sales_data[item_name][day] = {"sales": 0, "earnings": 0}

            daily_sales_data[item_name][day]["sales"] += order.quantity
            daily_sales_data[item_name][day]["earnings"] += order.total
            total_earnings_per_day[day] += order.total
            days_with_data.add(day)

        columns = ["Item", "Price(₹)"]
        for day in days:
            if day in days_with_data:
                columns.append(f"{day} Sales")
                columns.append(f"{day} Earnings")

        worksheet.append(columns)

        for item_name, sales_data in daily_sales_data.items():
            row = [item_name, sales_data["price"]]
            for day in days:
                if day in days_with_data:
                    if day in sales_data:
                        row.append(sales_data[day]["sales"])
                        row.append(sales_data[day]["earnings"])
                    else:
                        row.append(0)
                        row.append(0)
            worksheet.append(row)

        total_row = ["Total", ""]
        for day in days:
            if day in days_with_data:
                total_row.append("")
                total_row.append(total_earnings_per_day[day])
        worksheet.append(total_row)

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response[
            "Content-Disposition"
        ] = f"attachment; filename=daily_sales_report.xlsx"
        workbook.save(response)
        return response
    except Exception as e:
        return HttpResponseServerError(
            f"An error occurred while generating the report: {e}"
        )


def redirect_sales_report_page(request):
    return render(request, "reports/generate_sales_report.html")


@login_required
def generate_product_sales_report_all(request):
    try:
        if request.method == "GET":
            start_date = request.GET.get("start_date")
            end_date = request.GET.get("end_date")
            report_action = request.GET.get("report_action")

            if report_action == "monthly":
                return generate_product_sales_month_based_report_all(request)

            elif report_action == "daily":
                return generate_product_sales_day_based_report_all(request)

        return redirect_sales_report_page(request)
    except Exception as e:
        return HttpResponseServerError(
            f"An error occurred while generating the report: {e}"
        )
